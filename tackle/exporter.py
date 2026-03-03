import os
import re
import shutil
import tempfile
import time
import csv
from pathlib import Path
from typing import Iterable, List, Tuple, Optional, Dict, Any, Sequence

import pandas as pd

from .utils import _get_logger

logger = _get_logger(__name__)

def _excel_col_letter(col_num: int) -> str:
    """Convert a 1-based column number to an Excel column letter (e.g. 1->A, 27->AA)."""
    if col_num < 1:
        raise ValueError(f"col_num must be >= 1, got {col_num}")
    letters = ""
    n = int(col_num)
    while n:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters

def _excel_a1_range(nrows: int, ncols: int) -> Optional[str]:
    """Return an A1 range like A1:D10 for a rectangular region (best-effort)."""
    if nrows <= 0 or ncols <= 0:
        return None
    end_col = _excel_col_letter(int(ncols))
    return f"A1:{end_col}{int(nrows)}"

def _infer_col_widths(
    df: pd.DataFrame,
    *,
    max_rows: int = 200,
    min_width: int = 8,
    max_width: int = 60,
    kind: Optional[str] = None,
) -> Dict[str, int]:
    """Infer reasonable column widths (in approx characters) from headers + a small sample."""
    if df is None or df.empty:
        return {str(c): min(max_width, max(min_width, len(str(c)) + 2)) for c in (df.columns if df is not None else [])}

    sample = df.head(int(max_rows))
    widths: Dict[str, int] = {}

    def _clamp(value: int) -> int:
        return min(max_width, max(min_width, int(value)))

    overrides: Dict[str, int] = {}
    if kind in ("volcano", "volcano_merged"):
        overrides.update(
            {
                "GeneID": 12,
                "GeneSymbol": 16,
                "FunCats": 30,
                "GeneDescription": 50,
                "Description": 50,
            }
        )
    if kind == "sources":
        overrides.update(
            {
                "relative_path": 60,
                "sheet": 31,
                "kind": 12,
                "rows": 10,
                "cols": 10,
                "mtime": 20,
            }
        )

    for col in df.columns:
        col_name = str(col)
        if col_name in overrides:
            widths[col_name] = _clamp(overrides[col_name])
            continue

        # Prefer fixed widths for numeric-looking volcano metrics.
        metric = col_name.split("__", 1)[-1] if "__" in col_name else col_name
        if metric in ("pAdj", "pValue"):
            widths[col_name] = _clamp(14)
            continue
        if metric in ("log2_FC", "signedlogP", "t"):
            widths[col_name] = _clamp(12)
            continue

        header_len = len(col_name)
        try:
            ser = sample[col]
            # Avoid long float reprs: keep numeric columns to a sane width.
            if pd.api.types.is_numeric_dtype(ser):
                widths[col_name] = _clamp(max(header_len + 2, 12))
                continue
            max_val_len = int(ser.astype(str).map(len).max())
        except Exception:
            max_val_len = header_len
        widths[col_name] = _clamp(max(header_len, max_val_len) + 2)

    return widths

def _xlsxwriter_formats(workbook):
    """Create a small set of reusable xlsxwriter formats."""
    header = workbook.add_format(
        {
            "bold": True,
            "bg_color": "#F2F2F2",
            "border": 1,
            "text_wrap": False,
        }
    )
    text_wrap = workbook.add_format({"text_wrap": True})
    num_sci = workbook.add_format({"num_format": "0.00E+00"})
    num_dec3 = workbook.add_format({"num_format": "0.000"})
    num_int = workbook.add_format({"num_format": "0"})
    return {
        "header": header,
        "text_wrap": text_wrap,
        "num_sci": num_sci,
        "num_dec3": num_dec3,
        "num_int": num_int,
    }

def _xlsxwriter_col_format(header: str, *, kind: Optional[str], formats: Dict[str, Any]):
    """Return an xlsxwriter format for a column header (or None)."""
    name = str(header)
    metric = name.split("__", 1)[-1] if "__" in name else name
    if kind == "sources" and metric in ("rows", "cols"):
        return formats.get("num_int")
    if metric in ("pAdj", "pValue"):
        return formats.get("num_sci")
    if metric in ("log2_FC", "signedlogP", "t"):
        return formats.get("num_dec3")
    if metric in ("GeneDescription", "Description", "FunCats", "relative_path"):
        return formats.get("text_wrap")
    return None

def _apply_sheet_style_xlsxwriter(
    worksheet,
    *,
    nrows_total: int,
    ncols: int,
    headers: Sequence[str],
    kind: Optional[str],
    col_widths: Optional[Dict[str, int]],
    formats: Dict[str, Any],
) -> None:
    if nrows_total <= 0 or ncols <= 0:
        return
    last_row = max(0, int(nrows_total) - 1)
    last_col = max(0, int(ncols) - 1)
    worksheet.freeze_panes(1, 0)
    worksheet.autofilter(0, 0, last_row, last_col)
    worksheet.set_row(0, None, formats.get("header"))

    # Set widths / column formats (row format should keep header styling).
    for idx, h in enumerate(headers):
        hname = str(h)
        width = None
        if col_widths and hname in col_widths:
            width = col_widths[hname]
        if width is None:
            width = min(60, max(8, len(hname) + 2))
        fmt = _xlsxwriter_col_format(hname, kind=kind, formats=formats)
        worksheet.set_column(idx, idx, width, fmt)

def _apply_sheet_style_openpyxl(
    worksheet,
    *,
    nrows_total: int,
    ncols: int,
    headers: Sequence[str],
    kind: Optional[str],
    col_widths: Optional[Dict[str, int]],
) -> None:
    if nrows_total <= 0 or ncols <= 0:
        return
    try:
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return

    worksheet.freeze_panes = "A2"
    ref = _excel_a1_range(nrows_total, ncols)
    if ref:
        worksheet.auto_filter.ref = ref

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    for col_idx in range(1, int(ncols) + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    if col_widths:
        for col_idx, h in enumerate(headers, start=1):
            hname = str(h)
            width = col_widths.get(hname)
            if width is None:
                continue
            letter = _excel_col_letter(col_idx)
            worksheet.column_dimensions[letter].width = float(width)

def _read_tsv(path: Path, *, engine: str = "auto", low_memory: bool = False) -> pd.DataFrame:
    """Read a TSV file, optionally using pyarrow for speed."""
    if engine not in ("auto", "pandas", "pyarrow"):
        raise ValueError(f"Unsupported TSV engine: {engine}")

    if engine in ("auto", "pyarrow"):
        try:
            import pyarrow.csv as pv

            table = pv.read_csv(
                path,
                parse_options=pv.ParseOptions(delimiter="\t"),
            )
            return table.to_pandas()
        except Exception:
            if engine == "pyarrow":
                raise

    return pd.read_csv(path, sep="\t", low_memory=low_memory)

def _sanitize_sheet_name(name: str) -> str:
    # Excel sheet names: max 31 chars; cannot contain: : \ / ? * [ ]
    invalid = r'[:\\/?*\[\]]'
    safe = re.sub(invalid, "_", str(name))
    safe = safe.strip()
    # Avoid empty names
    if not safe:
        safe = "Sheet"
    # Excel also dislikes names that end with a single quote
    if safe.endswith("'"):
        safe = safe[:-1]
    if len(safe) > 31:
        safe = safe[:31]
    return safe

def _unique_sheet_name(base: str, existing: Sequence[str]) -> str:
    """Return a unique Excel sheet name within 31 chars, preserving a numeric suffix."""
    safe = _sanitize_sheet_name(base)
    if safe not in existing:
        return safe
    i = 2
    while True:
        suffix = f"_{i}"
        max_len = 31 - len(suffix)
        trimmed = safe[:max_len] if max_len > 0 else safe[:31]
        candidate = f"{trimmed}{suffix}"
        if candidate not in existing:
            return candidate
        i += 1


def _sanitize_label(value: str) -> str:
    """Sanitize a label for use as a column prefix without truncation.

    Removes characters that Excel sheet headers may not like but does not enforce
    a 31-character limit (which only applies to sheet names, not headers).
    """
    invalid = r'[:\\/?*\[\]]'
    safe = re.sub(invalid, "_", str(value))
    # Normalize whitespace to underscores
    safe = re.sub(r"\s+", "_", safe)
    return safe.strip("_")

def _stream_tsv_to_worksheet(
    path: Path,
    worksheet,
    *,
    max_width_rows: int = 200,
) -> Tuple[int, int, List[str], Optional[List[int]]]:
    """Stream a TSV into an xlsxwriter worksheet. Returns (rows, cols, header, widths)."""
    csv.field_size_limit(1024 * 1024 * 1024)
    rows = 0
    cols = 0
    header: List[str] = []
    max_lens: Optional[List[int]] = None
    with path.open("r", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for row_idx, row in enumerate(reader):
            if row_idx == 0:
                header = [str(x) for x in row]
                cols = len(row)
                max_lens = [len(str(x)) for x in row]
            worksheet.write_row(row_idx, 0, row)
            if max_lens is not None and row_idx <= int(max_width_rows):
                for i, val in enumerate(row):
                    if i >= len(max_lens):
                        break
                    try:
                        max_lens[i] = max(max_lens[i], len(str(val)))
                    except Exception:
                        continue
            rows += 1
    widths = None
    if max_lens:
        widths = [min(60, max(8, int(ml) + 2)) for ml in max_lens]
    return max(0, rows - 1), cols, header, widths

def _collect_tsvs(root: Path, patterns: Iterable[Tuple[str, str]]) -> List[Tuple[str, Path]]:
    logger.info("Exporter: scanning root=%s for patterns=%s", root, list(patterns))
    files: List[Tuple[str, Path]] = []
    for subdir, pat in patterns:
        # Prefer the canonical subdir under the provided root, but fall back to
        # recursively searching for any matching subdir deeper in the tree.
        search_dirs: List[Path] = []
        if subdir:
            d = root / subdir
            if d.exists():
                search_dirs = [d]
            else:
                # Recursively discover candidate subdirs named like 'export'/'volcano'
                search_dirs = [p for p in root.rglob(subdir) if p.is_dir()]
            logger.info(
                "Exporter: searching %d directories for '%s' files: %s",
                len(search_dirs), pat, ", ".join(str(p) for p in search_dirs) or "<none>"
            )
        else:
            search_dirs = [root]

        for d in search_dirs:
            for p in d.rglob(pat):
                try:
                    rel = p.relative_to(root)
                except ValueError:
                    # If p is outside root due to symlinks, fallback to basename
                    rel = Path(subdir) / p.name if subdir else p.name
                files.append((str(rel), p))
    logger.info("Exporter: discovered %d TSV files", len(files))
    # Stable order
    files.sort()
    return files


def build_export_xlsx(
    base_dir: str,
    out_path: str,
    include_export: bool = True,
    include_volcano: bool = True,
    pheno_df: Optional[pd.DataFrame] = None,
    meta: Optional[Dict[str, Any]] = None,
    *,
    engine_preference: Optional[str] = None,
    filter_contains: Optional[Sequence[str]] = None,
    slim_volcano: bool = True,
    volcano_topn: Optional[int] = None,
    merge_volcano: bool = False,
    staging_dir: Optional[str] = None,
    timing: bool = False,
    tsv_engine: str = "auto",
    pandas_low_memory: bool = False,
    stream_export: bool = False,
    style: bool = True,
) -> str:
    """
    Create an Excel workbook summarizing common tackle exports under an analysis directory.

    Parameters
    ----------
    base_dir: Root analysis output directory (e.g., data_obj.outpath)
    out_path: Destination .xlsx path
    include_export: Include tables from the 'export/' subdir
    include_volcano: Include volcano result TSVs
    """
    root = Path(base_dir)
    logger.info(
        "Exporter: build_export_xlsx base_dir=%s out_path=%s include_export=%s include_volcano=%s",
        root, out_path, include_export, include_volcano,
    )
    patterns: List[Tuple[str, str]] = []
    if include_export:
        patterns.append(("export", "*.tsv"))
    if include_volcano:
        patterns.append(("volcano", "*.tsv"))

    targets = _collect_tsvs(root, patterns)
    if filter_contains:
        fc = [s for s in filter_contains if s]
        if fc:
            logger.info("Exporter: filtering %d files by substrings: %s", len(targets), ", ".join(fc))
            targets = [(rel, p) for (rel, p) in targets if any(s in str(rel) for s in fc)]
    if not targets:
        logger.error("Exporter: no TSVs found under %s for patterns %s", base_dir, patterns)
        raise FileNotFoundError(f"No TSV exports found under {base_dir} matching \n{patterns}")

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    write_target = out
    staging_path: Optional[Path] = None
    if staging_dir:
        staging_root = Path(staging_dir)
        staging_root.mkdir(parents=True, exist_ok=True)
        fd, tmp_path = tempfile.mkstemp(prefix="tackle_xlsx_", suffix=".xlsx", dir=staging_root)
        os.close(fd)
        staging_path = Path(tmp_path)
        write_target = staging_path
        logger.info("Exporter: staging workbook at %s (final %s)", write_target, out)

    # Choose an engine if available
    # Prefer xlsxwriter for faster writes when available
    engine = None
    candidates = ("xlsxwriter", "openpyxl")
    if engine_preference:
        candidates = (engine_preference,) + tuple(c for c in candidates if c != engine_preference)
    for candidate in candidates:
        try:
            __import__(candidate)
            engine = candidate
            break
        except Exception:
            continue
    if engine is None:
        logger.error("Exporter: no Excel writer engine available (need openpyxl or xlsxwriter)")
        raise RuntimeError(
            "No Excel writer engine found. Install one of 'openpyxl' or 'xlsxwriter' to create XLSX exports."
        )
    logger.info("Exporter: using Excel engine '%s'", engine)
    engine_kwargs = None
    if engine == "xlsxwriter":
        engine_kwargs = {
            "options": {
                "strings_to_urls": False,
                "strings_to_formulas": False,
                "use_shared_strings": False,
            }
        }

    with pd.ExcelWriter(write_target, engine=engine, engine_kwargs=engine_kwargs) as xw:
        xlsx_formats = _xlsxwriter_formats(xw.book) if (style and engine == "xlsxwriter") else None
        sources_rows: List[Dict[str, Any]] = []
        # Manifest sheet removed per request; key metadata remains in logs/context
        # Optional phenotype/metadata sheet first

        if pheno_df is not None and isinstance(pheno_df, pd.DataFrame) and not pheno_df.empty:
            pheno = pheno_df.copy()
            pheno.index.name = pheno.index.name or "Sample"
            pheno_reset = pheno.reset_index()
            pheno_sheet = _sanitize_sheet_name("phenotype")
            logger.info("Exporter: adding sheet '%s' with shape %s", pheno_sheet, pheno_reset.shape)
            t0 = time.perf_counter()
            pheno_reset.to_excel(xw, sheet_name=pheno_sheet, index=False)
            if style:
                try:
                    headers = [str(c) for c in pheno_reset.columns]
                    widths = _infer_col_widths(pheno_reset, kind="phenotype")
                    ws = xw.sheets.get(pheno_sheet)
                    if ws is not None:
                        if engine == "xlsxwriter":
                            _apply_sheet_style_xlsxwriter(
                                ws,
                                nrows_total=int(pheno_reset.shape[0]) + 1,
                                ncols=int(pheno_reset.shape[1]),
                                headers=headers,
                                kind="phenotype",
                                col_widths=widths,
                                formats=xlsx_formats or {},
                            )
                        elif engine == "openpyxl":
                            _apply_sheet_style_openpyxl(
                                ws,
                                nrows_total=int(pheno_reset.shape[0]) + 1,
                                ncols=int(pheno_reset.shape[1]),
                                headers=headers,
                                kind="phenotype",
                                col_widths=widths,
                            )
                except Exception:
                    logger.debug("Exporter: styling skipped for sheet '%s'", pheno_sheet)
            if timing:
                logger.info("Exporter: wrote sheet '%s' in %.2fs", pheno_sheet, time.perf_counter() - t0)
        # Optionally merge all volcano TSVs into a single wide sheet
        volcano_targets = [(rel, p) for (rel, p) in targets if str(rel).startswith("volcano")]
        export_targets = [(rel, p) for (rel, p) in targets if str(rel).startswith("export")]

        volcano_wide_metrics: Optional[pd.DataFrame] = None

        def _label_from_rel(relpath: str) -> str:
            base = os.path.basename(relpath)
            name = re.sub(r"\.tsv$", "", base, flags=re.IGNORECASE)
            m = re.search(r"(?<=group)[_]?(.*)$", name)
            lbl = m.group(1) if m else name
            # Use a generous, sanitized label for column prefixes (no 31-char cap needed)
            return _sanitize_label(lbl)

        merged_volcano = False
        if merge_volcano and volcano_targets:
            logger.info("Exporter: merging %d volcano tables into a single wide sheet", len(volcano_targets))
            merged: Optional[pd.DataFrame] = None
            seen_labels = set()
            annot_df: Optional[pd.DataFrame] = None  # index = GeneID; columns: FunCats, GeneDescription
            keep_cols_base = [
                "GeneID",
                "GeneSymbol",
                "log2_FC",
                "pAdj",
                "pValue",
                "signedlogP",
                "t",
            ]
            for idx, (rel, path) in enumerate(volcano_targets, start=1):
                if timing:
                    logger.info("Exporter: reading volcano '%s'", rel)
                t_read = time.perf_counter()
                df = _read_tsv(path, engine=tsv_engine, low_memory=pandas_low_memory)
                read_elapsed = time.perf_counter() - t_read
                df.columns = [str(c) for c in df.columns]
                if "GeneID" not in df.columns:
                    logger.warning("Exporter: skipping volcano file without GeneID column: %s", rel)
                    continue
                df["GeneID"] = df["GeneID"].astype(str)
                # Capture gene annotations for front matter (once, coalescing across files)
                try:
                    cand = {}
                    if "FunCats" in df.columns:
                        cand["FunCats"] = df["FunCats"].astype(str)
                    if "GeneDescription" in df.columns:
                        cand["GeneDescription"] = df["GeneDescription"].astype(str)
                    elif "Description" in df.columns:
                        cand["GeneDescription"] = df["Description"].astype(str)
                    if cand:
                        ad = pd.DataFrame(cand)
                        ad.insert(0, "GeneID", df["GeneID"].astype(str))
                        ad = ad.drop_duplicates(subset=["GeneID"]).set_index("GeneID")
                        annot_df = ad if annot_df is None else annot_df.combine_first(ad)
                except Exception:
                    logger.debug("Exporter: unable to capture annotations from %s", rel)
                label = _label_from_rel(rel)
                # Ensure uniqueness
                if label in seen_labels:
                    i = 2
                    while f"{label}_{i}" in seen_labels:
                        i += 1
                    label = f"{label}_{i}"
                seen_labels.add(label)
                # Select columns
                cols = [c for c in keep_cols_base if c in df.columns]
                slim_df = df[cols] if slim_volcano and cols else df
                if slim_volcano and cols:
                    logger.info("Exporter: slim-volcano applied to '%s' (columns kept: %s)", rel, cols)
                # Optionally reduce to top-N rows for speed
                if volcano_topn:
                    if "pAdj" in slim_df.columns:
                        slim_df = slim_df.sort_values("pAdj").head(int(volcano_topn))
                    elif "pValue" in slim_df.columns:
                        slim_df = slim_df.sort_values("pValue").head(int(volcano_topn))
                    else:
                        slim_df = slim_df.head(int(volcano_topn))
                # Rename metric columns with label prefix
                rename_map = {}
                for c in slim_df.columns:
                    if c in ("GeneID", "GeneSymbol"):
                        continue
                    rename_map[c] = f"{label}__{c}"
                slim_df = slim_df.rename(columns=rename_map)
                # Merge outer on GeneID; preserve GeneSymbol if present
                if merged is None:
                    merged = slim_df.copy()
                else:
                    merged = pd.merge(merged, slim_df, on=["GeneID", *(["GeneSymbol"] if "GeneSymbol" in merged.columns and "GeneSymbol" in slim_df.columns else [])], how="outer")
                logger.info("Exporter: merged '%s' -> current shape %s", rel, merged.shape if merged is not None else None)
                if timing:
                    mem_mb = df.memory_usage(deep=True).sum() / 1e6
                    logger.info(
                        "Exporter: read '%s' in %.2fs (mem %.1f MB)",
                        rel,
                        read_elapsed,
                        mem_mb,
                    )
                # Track sources
                try:
                    stat = path.stat(); mtime = stat.st_mtime
                except Exception:
                    mtime = None
                sources_rows.append({
                    "sheet": "volcano_merged",
                    "relative_path": str(rel),
                    "rows": int(df.shape[0]),
                    "cols": int(df.shape[1]),
                    "mtime": mtime,
                    "kind": "volcano",
                })
            if merged is not None:
                # Attach front-matter annotations if available
                if annot_df is not None:
                    try:
                        merged = merged.merge(annot_df.reset_index(), on="GeneID", how="left")
                    except Exception:
                        logger.debug("Exporter: failed to merge annotations onto merged volcano table")
                # Order columns: front matter then metrics grouped by type
                front = [c for c in ("GeneID", "GeneSymbol", "FunCats", "GeneDescription") if c in merged.columns]
                # Discover metric columns with pattern '<label>__<metric>'
                metric_cols = [c for c in merged.columns if c not in front]
                label_metric = {}
                for c in metric_cols:
                    m = re.match(r"(.+?)__([A-Za-z0-9_.]+)$", c)
                    if not m:
                        continue
                    lbl, met = m.group(1), m.group(2)
                    label_metric.setdefault(met, []).append((lbl, c))
                # Desired metric ordering
                metric_order = ["log2_FC", "signedlogP", "pValue", "pAdj", "t"]
                ordered_metric_cols: List[str] = []
                for met in metric_order:
                    pairs = label_metric.get(met, [])
                    # sort by label for deterministic order
                    for _lbl, col in sorted(pairs, key=lambda x: x[0]):
                        ordered_metric_cols.append(col)
                # Append any leftover metric columns not matched above
                leftover = [c for c in metric_cols if c not in ordered_metric_cols]
                ordered_cols = front + ordered_metric_cols + leftover
                merged = merged[ordered_cols]
                try:
                    metric_cols = [c for c in merged.columns if c not in front and c != "GeneID"]
                    if metric_cols:
                        volcano_wide_metrics = merged[["GeneID", *metric_cols]].copy()
                        volcano_wide_metrics["GeneID"] = volcano_wide_metrics["GeneID"].astype(str)
                        volcano_wide_metrics = volcano_wide_metrics.drop_duplicates(subset=["GeneID"], keep="first")
                except Exception:
                    logger.debug("Exporter: failed to cache wide volcano metrics for stitching")
                sheet_name = _sanitize_sheet_name("volcano_merged")
                logger.info("Exporter: adding merged volcano sheet '%s' with shape %s", sheet_name, merged.shape)
                t0 = time.perf_counter()
                merged.to_excel(xw, sheet_name=sheet_name, index=False)
                merged_volcano = True
                if style:
                    try:
                        headers = [str(c) for c in merged.columns]
                        widths = _infer_col_widths(merged, kind="volcano_merged")
                        ws = xw.sheets.get(sheet_name)
                        if ws is not None:
                            if engine == "xlsxwriter":
                                _apply_sheet_style_xlsxwriter(
                                    ws,
                                    nrows_total=int(merged.shape[0]) + 1,
                                    ncols=int(merged.shape[1]),
                                    headers=headers,
                                    kind="volcano_merged",
                                    col_widths=widths,
                                    formats=xlsx_formats or {},
                                )
                            elif engine == "openpyxl":
                                _apply_sheet_style_openpyxl(
                                    ws,
                                    nrows_total=int(merged.shape[0]) + 1,
                                    ncols=int(merged.shape[1]),
                                    headers=headers,
                                    kind="volcano_merged",
                                    col_widths=widths,
                                )
                    except Exception:
                        logger.debug("Exporter: styling skipped for merged volcano sheet")
                if timing:
                    logger.info(
                        "Exporter: wrote sheet '%s' in %.2fs (rows %d, cols %d)",
                        sheet_name,
                        time.perf_counter() - t0,
                        merged.shape[0],
                        merged.shape[1],
                    )

        # Write non-volcano (export) tables as individual sheets, and volcano tables
        # individually only if not merging
        remaining = export_targets + ([] if merged_volcano else volcano_targets)
        total = len(remaining)
        for idx, (rel, path) in enumerate(remaining, start=1):
            # Deduce a concise sheet name from relative path
            base = str(rel)
            base = base.replace(os.sep, "__")
            base = re.sub(r"\.tsv$", "", base, flags=re.IGNORECASE)
            final = _unique_sheet_name(base, xw.sheets.keys())

            is_export = str(rel).startswith("export")
            can_stream = stream_export and engine == "xlsxwriter" and is_export
            if can_stream:
                if timing:
                    logger.info("Exporter: streaming '%s' -> sheet '%s'", rel, final)
                worksheet = xw.book.add_worksheet(final)
                xw.sheets[final] = worksheet
                t0 = time.perf_counter()
                rows, cols, headers, widths_list = _stream_tsv_to_worksheet(path, worksheet)
                logger.info(
                    "Exporter: [%d/%d] added sheet '%s' from '%s' with shape (%d, %d) (streamed)",
                    idx, total, final, rel, rows, cols,
                )
                if style:
                    try:
                        widths = None
                        if headers and widths_list and len(headers) == len(widths_list):
                            widths = {str(h): int(w) for h, w in zip(headers, widths_list)}
                        _apply_sheet_style_xlsxwriter(
                            worksheet,
                            nrows_total=int(rows) + 1,
                            ncols=int(cols),
                            headers=headers or [f"col{i+1}" for i in range(int(cols))],
                            kind="export",
                            col_widths=widths,
                            formats=xlsx_formats or {},
                        )
                    except Exception:
                        logger.debug("Exporter: styling skipped for streamed sheet '%s'", final)
                if timing:
                    logger.info(
                        "Exporter: wrote sheet '%s' in %.2fs (streamed)",
                        final,
                        time.perf_counter() - t0,
                    )
                try:
                    stat = path.stat()
                    mtime = stat.st_mtime
                except Exception:
                    mtime = None
                sources_rows.append({
                    "sheet": final,
                    "relative_path": str(rel),
                    "rows": int(rows),
                    "cols": int(cols),
                    "mtime": mtime,
                    "kind": "export",
                })
                continue

            if timing:
                logger.info("Exporter: reading '%s'", rel)
            t_read = time.perf_counter()
            df = _read_tsv(path, engine=tsv_engine, low_memory=pandas_low_memory)
            read_elapsed = time.perf_counter() - t_read
            orig_shape = df.shape

            # Stitch merged volcano metrics to MSPC export tables for convenience.
            try:
                if volcano_wide_metrics is not None and is_export:
                    base_name = os.path.basename(str(rel)).lower()
                    if "data_mspc_" in base_name and "geneid" in [str(c).lower() for c in df.columns]:
                        # Normalize join keys as strings
                        df.columns = [str(c) for c in df.columns]
                        df["GeneID"] = df["GeneID"].astype(str)
                        missing_cols = [
                            c
                            for c in volcano_wide_metrics.columns
                            if c != "GeneID" and c not in df.columns
                        ]
                        if missing_cols:
                            df = df.merge(
                                volcano_wide_metrics[["GeneID", *missing_cols]],
                                on="GeneID",
                                how="left",
                            )
            except Exception:
                logger.debug("Exporter: failed to stitch volcano metrics onto %s", rel)

            # Optional volcano slimming: keep only summary columns and/or top-N
            if slim_volcano and str(rel).startswith("volcano"):
                keep_cols = [
                    "GeneID",
                    "GeneSymbol",
                    "log2_FC",
                    "pAdj",
                    "pValue",
                    "signedlogP",
                    "t",
                ]
                cols = [c for c in keep_cols if c in df.columns]
                if cols:
                    df = df[cols]
            if volcano_topn and str(rel).startswith("volcano"):
                # If sorting columns exist, use them; otherwise just head()
                if "pAdj" in df.columns:
                    df = df.sort_values("pAdj", ascending=True).head(int(volcano_topn))
                elif "pValue" in df.columns:
                    df = df.sort_values("pValue", ascending=True).head(int(volcano_topn))
                else:
                    df = df.head(int(volcano_topn))
            logger.info(
                "Exporter: [%d/%d] adding sheet '%s' from '%s' with shape %s (orig %s)",
                idx, total, final, rel, df.shape, orig_shape,
            )
            t0 = time.perf_counter()
            df.to_excel(xw, sheet_name=final, index=False)
            if style:
                try:
                    headers = [str(c) for c in df.columns]
                    kind = "export" if str(rel).startswith("export") else ("volcano" if str(rel).startswith("volcano") else "other")
                    widths = _infer_col_widths(df, kind=kind)
                    ws = xw.sheets.get(final)
                    if ws is not None:
                        if engine == "xlsxwriter":
                            _apply_sheet_style_xlsxwriter(
                                ws,
                                nrows_total=int(df.shape[0]) + 1,
                                ncols=int(df.shape[1]),
                                headers=headers,
                                kind=kind,
                                col_widths=widths,
                                formats=xlsx_formats or {},
                            )
                        elif engine == "openpyxl":
                            _apply_sheet_style_openpyxl(
                                ws,
                                nrows_total=int(df.shape[0]) + 1,
                                ncols=int(df.shape[1]),
                                headers=headers,
                                kind=kind,
                                col_widths=widths,
                            )
                except Exception:
                    logger.debug("Exporter: styling skipped for sheet '%s'", final)
            if timing:
                mem_mb = df.memory_usage(deep=True).sum() / 1e6
                logger.info(
                    "Exporter: wrote sheet '%s' in %.2fs (read %.2fs, mem %.1f MB)",
                    final,
                    time.perf_counter() - t0,
                    read_elapsed,
                    mem_mb,
                )

            try:
                stat = path.stat()
                mtime = stat.st_mtime
            except Exception:
                mtime = None
            sources_rows.append({
                "sheet": final,
                "relative_path": str(rel),
                "rows": int(df.shape[0]),
                "cols": int(df.shape[1]),
                "mtime": mtime,
                "kind": "volcano" if str(rel).startswith("volcano") else ("export" if str(rel).startswith("export") else "other"),
            })

        # Add a 'sources' sheet last with a summary of included TSVs
        try:
            if sources_rows:
                sources_df = pd.DataFrame(sources_rows)
                logger.info("Exporter: adding 'sources' sheet with %d entries", len(sources_df))
                sources_df.to_excel(xw, sheet_name=_sanitize_sheet_name("sources"), index=False)
                if style:
                    try:
                        headers = [str(c) for c in sources_df.columns]
                        widths = _infer_col_widths(sources_df, kind="sources")
                        sheet_name = _sanitize_sheet_name("sources")
                        ws = xw.sheets.get(sheet_name)
                        if ws is not None:
                            if engine == "xlsxwriter":
                                _apply_sheet_style_xlsxwriter(
                                    ws,
                                    nrows_total=int(sources_df.shape[0]) + 1,
                                    ncols=int(sources_df.shape[1]),
                                    headers=headers,
                                    kind="sources",
                                    col_widths=widths,
                                    formats=xlsx_formats or {},
                                )
                            elif engine == "openpyxl":
                                _apply_sheet_style_openpyxl(
                                    ws,
                                    nrows_total=int(sources_df.shape[0]) + 1,
                                    ncols=int(sources_df.shape[1]),
                                    headers=headers,
                                    kind="sources",
                                    col_widths=widths,
                                )
                    except Exception:
                        logger.debug("Exporter: styling skipped for 'sources' sheet")
        except Exception:
            logger.exception("Exporter: failed to add sources sheet")

    if staging_path is not None:
        shutil.move(str(staging_path), str(out))
    logger.info("Exporter: wrote workbook %s", out)
    return str(out)
