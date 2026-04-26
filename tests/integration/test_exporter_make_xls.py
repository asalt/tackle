import json
import os
from pathlib import Path

import pandas as pd
import pytest

from tackle.exporter import build_export_xlsx, build_xlsx_visual_review_bundle


def _has_excel_writer():
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:
        try:
            import xlsxwriter  # noqa: F401
            # Writer exists, but reading sheets will still require openpyxl
            return True
        except Exception:
            return False


def _has_soffice():
    try:
        import shutil
        return bool(shutil.which("soffice") or shutil.which("libreoffice"))
    except Exception:
        return False


def _has_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:
        return False


def _load_sheetnames(path: str):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        return [s.lower() for s in wb.sheetnames]
    except Exception:
        return None


@pytest.mark.skipif(not _has_excel_writer(), reason="No Excel writer engine available")
def test_build_export_xlsx_writes_pheno_and_sources(tmp_path):
    base = tmp_path
    (base / "export").mkdir(parents=True, exist_ok=True)
    (base / "volcano").mkdir(parents=True, exist_ok=True)

    # Minimal export and volcano TSVs
    pd.DataFrame({"GeneID": ["g1", "g2"], "Val": [1.0, 2.0]}).to_csv(
        base / "export" / "data_area.tsv", sep="\t", index=False
    )
    pd.DataFrame({
        "GeneID": ["g1", "g2"],
        "log2_FC": [1.2, -0.8],
        "pAdj": [0.01, 0.2],
        "pValue": [0.01, 0.2],
    }).to_csv(base / "volcano" / "condA_by_condB.tsv", sep="\t", index=False)

    pheno = pd.DataFrame({"condition": ["A", "A", "B", "B"]}, index=["s1", "s2", "s3", "s4"])\
             .rename_axis("Sample")

    out = base / "export" / "summary.xlsx"
    result = build_export_xlsx(
        base_dir=str(base),
        out_path=str(out),
        include_export=True,
        include_volcano=True,
        pheno_df=pheno,
        meta={"analysis_name": "TEST"},
    )
    assert os.path.exists(result)

    sheets = _load_sheetnames(result)
    if sheets is not None:
        assert "phenotype" in sheets
        assert "sources" in sheets
        assert any(s.startswith("export__") for s in sheets)
        assert any(s.startswith("volcano__") for s in sheets)


@pytest.mark.skipif(not _has_excel_writer(), reason="No Excel writer engine available")
def test_build_export_xlsx_respects_include_flags(tmp_path):
    base = tmp_path
    (base / "export").mkdir(parents=True, exist_ok=True)
    (base / "volcano").mkdir(parents=True, exist_ok=True)

    pd.DataFrame({"GeneID": ["g1"], "Val": [1.0]}).to_csv(
        base / "export" / "data_area.tsv", sep="\t", index=False
    )
    pd.DataFrame({"GeneID": ["g1"], "log2_FC": [2.0], "pAdj": [0.01], "pValue": [0.01]}).to_csv(
        base / "volcano" / "one.tsv", sep="\t", index=False
    )

    pheno = pd.DataFrame(index=["s1"]).rename_axis("Sample")

    out1 = base / "export" / "summary.noexport.xlsx"
    res1 = build_export_xlsx(
        base_dir=str(base),
        out_path=str(out1),
        include_export=False,
        include_volcano=True,
        pheno_df=pheno,
        meta={"analysis_name": "TEST"},
    )
    sheets1 = _load_sheetnames(res1)
    if sheets1 is not None:
        assert not any(s.startswith("export__") for s in sheets1)
        assert any(s.startswith("volcano__") for s in sheets1)

    out2 = base / "export" / "summary.novolcano.xlsx"
    res2 = build_export_xlsx(
        base_dir=str(base),
        out_path=str(out2),
        include_export=True,
        include_volcano=False,
        pheno_df=pheno,
        meta={"analysis_name": "TEST"},
    )
    sheets2 = _load_sheetnames(res2)
    if sheets2 is not None:
        assert any(s.startswith("export__") for s in sheets2)
        assert not any(s.startswith("volcano__") for s in sheets2)


@pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl required to inspect workbook styling")
def test_build_export_xlsx_applies_minimal_styling_by_default(tmp_path):
    base = tmp_path
    (base / "export").mkdir(parents=True, exist_ok=True)
    (base / "volcano").mkdir(parents=True, exist_ok=True)

    pd.DataFrame({"GeneID": ["g1", "g2"], "Val": [1.0, 2.0]}).to_csv(
        base / "export" / "data_area.tsv", sep="\t", index=False
    )
    pd.DataFrame({
        "GeneID": ["g1", "g2"],
        "log2_FC": [1.2, -0.8],
        "pAdj": [0.01, 0.2],
        "pValue": [0.01, 0.2],
    }).to_csv(base / "volcano" / "condA_by_condB.tsv", sep="\t", index=False)

    out = base / "export" / "summary.xlsx"
    result = build_export_xlsx(
        base_dir=str(base),
        out_path=str(out),
        include_export=True,
        include_volcano=True,
        pheno_df=None,
        meta={"analysis_name": "TEST"},
    )
    assert os.path.exists(result)

    import openpyxl

    wb = openpyxl.load_workbook(result, read_only=False)
    export_sheet = next((s for s in wb.sheetnames if s.lower().startswith("export__")), None)
    assert export_sheet is not None
    ws = wb[export_sheet]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref


@pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl required to inspect workbook styling")
def test_build_export_xlsx_can_disable_styling(tmp_path):
    base = tmp_path
    (base / "export").mkdir(parents=True, exist_ok=True)
    pd.DataFrame({"GeneID": ["g1", "g2"], "Val": [1.0, 2.0]}).to_csv(
        base / "export" / "data_area.tsv", sep="\t", index=False
    )

    out = base / "export" / "summary.nostyle.xlsx"
    result = build_export_xlsx(
        base_dir=str(base),
        out_path=str(out),
        include_export=True,
        include_volcano=False,
        pheno_df=None,
        meta={"analysis_name": "TEST"},
        style=False,
    )
    assert os.path.exists(result)

    import openpyxl

    wb = openpyxl.load_workbook(result, read_only=False)
    export_sheet = next((s for s in wb.sheetnames if s.lower().startswith("export__")), None)
    assert export_sheet is not None
    ws = wb[export_sheet]
    assert ws.freeze_panes is None
    assert not ws.auto_filter.ref


@pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl required to inspect workbook contents")
def test_build_export_xlsx_stitches_volcano_to_mspc_export(tmp_path):
    base = tmp_path
    (base / "export").mkdir(parents=True, exist_ok=True)
    (base / "volcano").mkdir(parents=True, exist_ok=True)

    pd.DataFrame(
        {
            "GeneID": ["g1", "g2"],
            "GeneSymbol": ["A", "B"],
            "iBAQ_dstrAdj": [1.0, 2.0],
        }
    ).to_csv(base / "export" / "data_MSPC_iBAQ_dstrAdj.tsv", sep="\t", index=False)

    pd.DataFrame(
        {
            "GeneID": ["g1", "g2"],
            "log2_FC": [1.2, -0.8],
            "pAdj": [0.01, 0.2],
            "pValue": [0.01, 0.2],
            "t": [2.0, -1.1],
            "signedlogP": [2.0, -0.7],
        }
    ).to_csv(base / "volcano" / "condA_by_condB.tsv", sep="\t", index=False)

    out = base / "export" / "summary.xlsx"
    result = build_export_xlsx(
        base_dir=str(base),
        out_path=str(out),
        include_export=True,
        include_volcano=True,
        pheno_df=None,
        meta={"analysis_name": "TEST"},
        merge_volcano=True,
    )
    assert os.path.exists(result)

    import openpyxl

    wb = openpyxl.load_workbook(result, read_only=False)
    # export__data_MSPC_iBAQ_dstrAdj
    export_sheet = next((s for s in wb.sheetnames if s.lower().startswith("export__data_mspc_ibaq_dstradj")), None)
    assert export_sheet is not None

    ws = wb[export_sheet]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    # Confirm at least one stitched volcano metric column exists and is appended (after base export cols).
    stitched_col = "condA_by_condB__log2_FC"
    assert stitched_col in headers
    assert headers.index(stitched_col) > headers.index("iBAQ_dstrAdj")


@pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl required to inspect workbook styling")
def test_build_export_xlsx_semantic_style_check_for_volcano_sheet(tmp_path):
    base = tmp_path
    (base / "volcano").mkdir(parents=True, exist_ok=True)

    pd.DataFrame(
        {
            "GeneID": ["g1", "g2"],
            "GeneSymbol": ["AAA", "BBB"],
            "GeneDescription": ["alpha protein", "beta protein"],
            "log2_FC": [1.2, -0.8],
            "pAdj": [0.01, 0.2],
        }
    ).to_csv(base / "volcano" / "condA_by_condB.tsv", sep="	", index=False)

    out = base / "export" / "summary.semantic.xlsx"
    result = build_export_xlsx(
        base_dir=str(base),
        out_path=str(out),
        include_export=False,
        include_volcano=True,
        pheno_df=None,
        meta={"analysis_name": "TEST"},
        engine_preference="openpyxl",
        slim_volcano=False,
    )

    import openpyxl

    wb = openpyxl.load_workbook(result, read_only=False)
    sheet_name = next((s for s in wb.sheetnames if s.lower().startswith("volcano__")), None)
    assert sheet_name is not None

    ws = wb[sheet_name]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:E3"
    assert ws["A1"].font.bold is True
    assert str(ws["A1"].fill.fgColor.rgb).endswith("F2F2F2")
    assert ws.column_dimensions["A"].width == pytest.approx(12.0)
    assert ws.column_dimensions["B"].width == pytest.approx(16.0)
    assert ws.column_dimensions["C"].width == pytest.approx(50.0)
    assert ws.column_dimensions["D"].width == pytest.approx(12.0)
    assert ws.column_dimensions["E"].width == pytest.approx(14.0)


@pytest.mark.skipif(not _has_soffice(), reason="soffice/libreoffice required for visual-check preview")
def test_build_export_xlsx_visual_check_renders_review_bundle_when_renderer_available(tmp_path):
    base = tmp_path
    (base / "export").mkdir(parents=True, exist_ok=True)

    pd.DataFrame(
        {
            "GeneID": ["g1", "g2"],
            "GeneDescription": ["alpha protein with a longer description", "beta protein with a longer description"],
            "VeryLongSummaryColumn": ["condition_A_reference", "condition_B_reference"],
            "Val": [1.0, 2.0],
        }
    ).to_csv(base / "export" / "data_area.tsv", sep="	", index=False)

    out = base / "export" / "summary.visual.xlsx"
    result = build_export_xlsx(
        base_dir=str(base),
        out_path=str(out),
        include_export=True,
        include_volcano=False,
        pheno_df=None,
        meta={"analysis_name": "TEST"},
        engine_preference="openpyxl",
    )

    bundle = build_xlsx_visual_review_bundle(
        result,
        fixture_name="wide layout visual",
        checklist=[
            "Check whether long headers and descriptions remain readable.",
            "Check whether the overall worksheet layout is reviewable from the rendered PNG.",
        ],
        out_dir=str(tmp_path / "visual-check"),
    )

    manifest = json.loads(Path(bundle.manifest_path).read_text(encoding="utf-8"))
    prompt = Path(bundle.review_prompt_path).read_text(encoding="utf-8")

    assert bundle.available is True
    assert bundle.rendered is True, bundle.reason
    assert bundle.png_paths
    assert all(Path(p).exists() and Path(p).stat().st_size > 0 for p in bundle.png_paths)
    assert manifest["fixture_name"] == "wide_layout_visual"
    assert manifest["rendered"] is True
    assert manifest["png_paths"] == bundle.png_paths
    assert "Check whether long headers and descriptions remain readable." in prompt
