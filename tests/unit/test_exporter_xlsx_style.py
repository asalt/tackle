import json
from pathlib import Path

import pandas as pd
import pytest

from tackle.exporter import (
    XlsxVisualCheckResult,
    XLSX_HEADER_FILL,
    XLSX_HEADER_FONT_NAME,
    _apply_sheet_style_openpyxl,
    _infer_col_widths,
    _xlsx_header_row_height,
    build_xlsx_visual_review_bundle,
    render_xlsx_visual_check,
)


def test_infer_col_widths_uses_stable_volcano_overrides():
    df = pd.DataFrame(
        {
            "GeneID": ["101"],
            "GeneSymbol": ["ABC"],
            "GeneDescription": ["alpha beta gamma"],
            "log2_FC": [1.234],
            "pAdj": [0.001],
        }
    )

    widths = _infer_col_widths(df, kind="volcano")

    assert widths["GeneID"] == 12
    assert widths["GeneSymbol"] == 16
    assert widths["GeneDescription"] == 50
    assert widths["log2_FC"] == 12
    assert widths["pAdj"] == 14


@pytest.mark.skipif(
    pytest.importorskip("openpyxl", reason="openpyxl required for workbook style checks") is None,
    reason="openpyxl required for workbook style checks",
)
def test_apply_sheet_style_openpyxl_sets_header_and_widths():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["GeneID", "GeneDescription", "pAdj"]
    ws.append(headers)
    ws.append(["101", "alpha protein", 0.001])

    _apply_sheet_style_openpyxl(
        ws,
        nrows_total=2,
        ncols=3,
        headers=headers,
        kind="volcano",
        col_widths={"GeneID": 12, "GeneDescription": 50, "pAdj": 14},
        header_rotation=60,
    )

    assert ws.freeze_panes == "C2"
    assert ws.auto_filter.ref == "A1:C2"
    assert ws.page_setup.orientation == "landscape"
    assert ws.sheet_view.showGridLines is False
    assert ws["A1"].font.bold is True
    assert ws["A1"].font.name == XLSX_HEADER_FONT_NAME
    assert ws["A1"].font.sz == 10
    assert str(ws["A1"].fill.fgColor.rgb).endswith(XLSX_HEADER_FILL)
    assert ws["A1"].alignment.text_rotation == 60
    assert ws["A1"].alignment.wrap_text is True
    assert ws.row_dimensions[1].height == pytest.approx(_xlsx_header_row_height(60))
    assert ws.column_dimensions["A"].width == pytest.approx(12.0)
    assert ws.column_dimensions["B"].width == pytest.approx(50.0)
    assert ws.column_dimensions["C"].width == pytest.approx(14.0)


@pytest.mark.skipif(
    pytest.importorskip("openpyxl", reason="openpyxl required for workbook style checks") is None,
    reason="openpyxl required for workbook style checks",
)
def test_apply_sheet_style_openpyxl_can_keep_horizontal_headers():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["GeneID", "GeneDescription"]
    ws.append(headers)
    ws.append(["101", "alpha protein"])

    _apply_sheet_style_openpyxl(
        ws,
        nrows_total=2,
        ncols=2,
        headers=headers,
        kind="volcano",
        col_widths={"GeneID": 12, "GeneDescription": 50},
        header_rotation=0,
    )

    assert ws["A1"].alignment.text_rotation == 0
    assert ws.row_dimensions[1].height == pytest.approx(_xlsx_header_row_height(0))


@pytest.mark.skipif(
    pytest.importorskip("openpyxl", reason="openpyxl required for workbook style checks") is None,
    reason="openpyxl required for workbook style checks",
)
def test_apply_sheet_style_openpyxl_can_color_scale_ibaq_columns():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["GeneID", "iBAQ_dstrAdj", "Other"]
    ws.append(headers)
    ws.append(["101", 1.0, 3.0])
    ws.append(["202", 10.0, 4.0])

    _apply_sheet_style_openpyxl(
        ws,
        nrows_total=3,
        ncols=3,
        headers=headers,
        kind="export",
        col_widths={"GeneID": 12, "iBAQ_dstrAdj": 14, "Other": 10},
        header_rotation=60,
        color_scale_ibaq=True,
    )

    ranges = [str(item.sqref) for item in ws.conditional_formatting]
    assert ranges == ["B2:B3"]


def test_render_xlsx_visual_check_reports_missing_renderer(tmp_path, monkeypatch):
    xlsx_path = tmp_path / "dummy.xlsx"
    xlsx_path.write_bytes(b"placeholder")

    monkeypatch.setattr("tackle.exporter.shutil.which", lambda _name: None)

    result = render_xlsx_visual_check(str(xlsx_path), out_dir=str(tmp_path / "preview"))

    assert result.available is False
    assert result.rendered is False
    assert result.output_paths == []
    assert "renderer" in str(result.reason).lower()


def test_build_xlsx_visual_review_bundle_writes_manifest_and_prompt(tmp_path, monkeypatch):
    xlsx_path = tmp_path / "visual.xlsx"
    xlsx_path.write_bytes(b"placeholder")
    rendered_png = tmp_path / "rendered.png"
    rendered_png.write_bytes(b"png")

    monkeypatch.setattr(
        "tackle.exporter.render_xlsx_visual_check",
        lambda *args, **kwargs: XlsxVisualCheckResult(
            available=True,
            rendered=True,
            output_paths=[str(rendered_png)],
            command=["soffice", "--headless"],
        ),
    )

    bundle = build_xlsx_visual_review_bundle(
        str(xlsx_path),
        fixture_name="volcano summary",
        checklist=["Check header readability", "Check width balance"],
        out_dir=str(tmp_path / "bundle"),
    )

    manifest = json.loads(Path(bundle.manifest_path).read_text(encoding="utf-8"))
    prompt = Path(bundle.review_prompt_path).read_text(encoding="utf-8")

    assert bundle.fixture_name == "volcano_summary"
    assert Path(bundle.manifest_path).exists()
    assert Path(bundle.review_prompt_path).exists()
    assert manifest["fixture_name"] == "volcano_summary"
    assert manifest["png_paths"] == [str(rendered_png)]
    assert manifest["checklist"] == ["Check header readability", "Check width balance"]
    assert manifest["review_status"] is None
    assert "Check header readability" in prompt
    assert "human or AI reviewer" in prompt


def test_build_xlsx_visual_review_bundle_uses_env_output_dir(tmp_path, monkeypatch):
    xlsx_path = tmp_path / "visual.xlsx"
    xlsx_path.write_bytes(b"placeholder")
    out_root = tmp_path / "kept-artifacts"

    monkeypatch.setenv("TACKLE_XLS_VISUAL_OUTDIR", str(out_root))
    monkeypatch.setattr(
        "tackle.exporter.render_xlsx_visual_check",
        lambda *args, **kwargs: XlsxVisualCheckResult(
            available=False,
            rendered=False,
            output_paths=[],
            command=[],
            reason="no renderer",
        ),
    )

    bundle = build_xlsx_visual_review_bundle(
        str(xlsx_path),
        fixture_name="basic export",
    )

    assert Path(bundle.bundle_dir).parent == out_root.resolve()
    assert Path(bundle.manifest_path).exists()


def test_build_xlsx_visual_review_bundle_preserves_failed_render_metadata(tmp_path, monkeypatch):
    xlsx_path = tmp_path / "visual.xlsx"
    xlsx_path.write_bytes(b"placeholder")

    monkeypatch.setattr(
        "tackle.exporter.render_xlsx_visual_check",
        lambda *args, **kwargs: XlsxVisualCheckResult(
            available=True,
            rendered=False,
            output_paths=[],
            command=["soffice"],
            reason="Renderer exited with code 1",
        ),
    )

    bundle = build_xlsx_visual_review_bundle(
        str(xlsx_path),
        fixture_name="wide layout visual",
        out_dir=str(tmp_path / "bundle"),
    )

    manifest = json.loads(Path(bundle.manifest_path).read_text(encoding="utf-8"))
    prompt = Path(bundle.review_prompt_path).read_text(encoding="utf-8")

    assert manifest["rendered"] is False
    assert manifest["reason"] == "Renderer exited with code 1"
    assert "Renderer exited with code 1" in prompt
